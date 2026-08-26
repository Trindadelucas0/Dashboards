from __future__ import annotations

import re
import time
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path

from bs4 import BeautifulSoup

OLE_MAGIC = b"\xd0\xcf\x11\xe0"
ZIP_MAGIC = b"PK"


@dataclass
class WorkbookGrid:
    path: str
    sheet_name: str
    rows: list[list[str]] = field(default_factory=list)
    kind: str = "unknown"

    def cell(self, r: int, c: int) -> str:
        if r < 0 or r >= len(self.rows):
            return ""
        row = self.rows[r]
        if c < 0 or c >= len(row):
            return ""
        return str(row[c] or "").strip()

    def row(self, r: int) -> list[str]:
        if r < 0 or r >= len(self.rows):
            return []
        return [str(x or "").strip() for x in self.rows[r]]


def is_placeholder_bytes(data: bytes) -> bool:
    """OneDrive/cópia falha: arquivo presente no disco, conteúdo zerado."""
    if not data:
        return True
    return not any(data[:2048])


def sniff(data: bytes) -> str:
    if is_placeholder_bytes(data):
        return "empty"
    head = data[:16]
    if head.startswith(OLE_MAGIC):
        return "xls"
    if head.startswith(ZIP_MAGIC):
        return "xlsx"
    sample = data[:4000].lstrip().lower()
    if sample.startswith(b"<html") or sample.startswith(b"<!doctype") or b"<table" in sample:
        return "html"
    if b"<?xml" in sample[:200]:
        return "xml"
    return "unknown"


def _html_to_grid(html: str) -> list[list[str]]:
    soup = BeautifulSoup(html, "lxml")
    tables = soup.find_all("table")
    if not tables:
        return []
    best = max(tables, key=lambda t: len(t.find_all("tr")))
    grid: list[list[str]] = []
    for tr in best.find_all("tr"):
        cells = tr.find_all(["td", "th"])
        grid.append([c.get_text(" ", strip=True) for c in cells])
    return grid


def safe_unlink(path: str | Path, attempts: int = 5) -> None:
    """Apaga arquivo temporário sem derrubar o preview se o Windows ainda o estiver usando."""
    p = Path(path)
    for i in range(attempts):
        try:
            p.unlink(missing_ok=True)
            return
        except OSError:
            time.sleep(0.12 * (i + 1))
    try:
        p.unlink(missing_ok=True)
    except OSError:
        pass


def _xlrd_sheet_rows(sheet) -> list[list[str]]:
    rows = []
    for r in range(sheet.nrows):
        row = []
        for c in range(sheet.ncols):
            val = sheet.cell_value(r, c)
            if isinstance(val, float) and val == int(val):
                row.append(str(int(val)))
            else:
                row.append(str(val).strip() if val is not None else "")
        rows.append(row)
    return rows


def _xlrd_grid(path: Path, sheet_index: int = 0) -> tuple[str, list[list[str]]]:
    import xlrd

    book = xlrd.open_workbook(str(path), formatting_info=False)
    try:
        sheet = book.sheet_by_index(sheet_index)
        return sheet.name, _xlrd_sheet_rows(sheet)
    finally:
        book.release_resources()


def _xlrd_all_grids(path: Path) -> list[tuple[str, list[list[str]]]]:
    import xlrd

    book = xlrd.open_workbook(str(path), formatting_info=False)
    try:
        return [(book.sheet_by_index(i).name, _xlrd_sheet_rows(book.sheet_by_index(i))) for i in range(book.nsheets)]
    finally:
        book.release_resources()


def _openpyxl_sheet_rows(ws) -> list[list[str]]:
    return [["" if v is None else str(v).strip() for v in row] for row in ws.iter_rows(values_only=True)]


def _openpyxl_grid(path: Path, data: bytes | None = None, sheet_index: int = 0) -> tuple[str, list[list[str]]]:
    from openpyxl import load_workbook

    src = BytesIO(data if data is not None else path.read_bytes())
    wb = load_workbook(src, data_only=True, read_only=True)
    try:
        ws = wb.worksheets[sheet_index]
        return ws.title, _openpyxl_sheet_rows(ws)
    finally:
        wb.close()
        src.close()


def _openpyxl_all_grids(path: Path, data: bytes | None = None) -> list[tuple[str, list[list[str]]]]:
    from openpyxl import load_workbook

    src = BytesIO(data if data is not None else path.read_bytes())
    wb = load_workbook(src, data_only=True, read_only=True)
    try:
        return [(ws.title, _openpyxl_sheet_rows(ws)) for ws in wb.worksheets]
    finally:
        wb.close()
        src.close()


def _cell_to_str(val) -> str:
    if val is None:
        return ""
    if isinstance(val, float):
        if val == int(val):
            return str(int(val))
        return str(val)
    if hasattr(val, "strftime"):
        try:
            return val.strftime("%d/%m/%Y")
        except Exception:  # noqa: BLE001
            return str(val)
    return str(val).strip()


_excel_lock = None
_excel_app = None


def _excel():
    global _excel_lock, _excel_app
    import threading
    import win32com.client  # type: ignore

    if _excel_lock is None:
        _excel_lock = threading.Lock()
    with _excel_lock:
        try:
            if _excel_app is not None:
                _excel_app.Visible
                return _excel_app
        except Exception:  # noqa: BLE001
            _excel_app = None
        app = win32com.client.Dispatch("Excel.Application")
        app.Visible = False
        app.DisplayAlerts = False
        app.ScreenUpdating = False
        _excel_app = app
        return _excel_app


def _com_used_range_to_rows(raw) -> list[list[str]]:
    rows: list[list[str]] = []
    if raw is None:
        return rows
    if not isinstance(raw, tuple):
        return [[_cell_to_str(raw)]]
    for row in raw:
        if not isinstance(row, tuple):
            rows.append([_cell_to_str(row)])
        else:
            rows.append([_cell_to_str(c) for c in row])
    return rows


def _com_grid(path: Path, sheet_index: int = 1) -> tuple[str, list[list[str]]]:
    global _excel_app
    last: Exception | None = None
    for attempt in range(2):
        try:
            excel = _excel()
            wb = excel.Workbooks.Open(str(path.resolve()), ReadOnly=True)
            try:
                ws = wb.Worksheets(sheet_index)
                rows = _com_used_range_to_rows(ws.UsedRange.Value)
                return str(ws.Name), rows
            finally:
                wb.Close(False)
        except Exception as exc:  # noqa: BLE001
            last = exc
            _excel_app = None
            if attempt == 0:
                time.sleep(0.4)
    raise last or RuntimeError("Excel COM falhou")


def _com_all_grids(path: Path) -> list[tuple[str, list[list[str]]]]:
    global _excel_app
    last: Exception | None = None
    for attempt in range(2):
        try:
            excel = _excel()
            wb = excel.Workbooks.Open(str(path.resolve()), ReadOnly=True)
            try:
                out: list[tuple[str, list[list[str]]]] = []
                for i in range(1, wb.Worksheets.Count + 1):
                    ws = wb.Worksheets(i)
                    out.append((str(ws.Name), _com_used_range_to_rows(ws.UsedRange.Value)))
                return out
            finally:
                wb.Close(False)
        except Exception as exc:  # noqa: BLE001
            last = exc
            _excel_app = None
            if attempt == 0:
                time.sleep(0.4)
    raise last or RuntimeError("Excel COM falhou")


def _load_html_grid(path: Path, data: bytes) -> WorkbookGrid:
    text = data.decode("latin-1", errors="ignore")
    rows = _html_to_grid(text)
    sheet = "HTML"
    title_m = re.search(r"<title>([^<]+)</title>", text, re.I)
    if title_m:
        sheet = title_m.group(1).strip()[:80]
    return WorkbookGrid(str(path), sheet, rows, "html")


def load_all_sheets(path: str | Path, data: bytes | None = None) -> list[WorkbookGrid]:
    """Carrega todas as abas (PIS+COFINS no mesmo .xls). HTML retorna uma grid só."""
    path = Path(path)
    if data is None:
        data = path.read_bytes()
    kind = sniff(data)
    errors: list[str] = []

    if kind == "empty":
        raise RuntimeError(f"{path.name}: arquivo vazio ou não baixado")

    if kind == "html":
        return [_load_html_grid(path, data)]

    if kind == "xlsx":
        try:
            pairs = _openpyxl_all_grids(path, data)
            return [WorkbookGrid(str(path), name, rows, "xlsx") for name, rows in pairs]
        except Exception as exc:  # noqa: BLE001
            errors.append(f"openpyxl:{exc}")

    if kind in ("xls", "unknown"):
        try:
            pairs = _xlrd_all_grids(path)
            return [WorkbookGrid(str(path), name, rows, "xls") for name, rows in pairs]
        except Exception as exc:  # noqa: BLE001
            errors.append(f"xlrd:{exc}")
        try:
            pairs = _com_all_grids(path)
            return [WorkbookGrid(str(path), name, rows, "com") for name, rows in pairs]
        except Exception as exc:  # noqa: BLE001
            errors.append(f"com:{exc}")
        try:
            text = data.decode("latin-1", errors="ignore")
            if "<table" in text.lower():
                return [_load_html_grid(path, data)]
        except Exception as exc:  # noqa: BLE001
            errors.append(f"html-fallback:{exc}")

    raise RuntimeError(f"Não foi possível ler {path.name}: {'; '.join(errors)}")


def load_workbook(path: str | Path, data: bytes | None = None) -> WorkbookGrid:
    sheets = load_all_sheets(path, data)
    return sheets[0]
